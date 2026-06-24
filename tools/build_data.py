#!/usr/bin/env python3
"""
build_data.py
=============
Génère `l18n/lqa-data.json` à partir de `lqa-textes-i18n.xlsx` (classeur i18n,
source de vérité du texte candidat), et détecte les clés référencées dans
`index.html` qui n'existent pas (encore) dans le classeur — Volet 3.

Usage :
    python build_data.py
    python build_data.py --xlsx chemin/vers/lqa-textes-i18n.xlsx \\
                          --html chemin/vers/index_Clau_0-5.html \\
                          --out  l18n/lqa-data.json \\
                          --missing-out l18n/lqa-cles-manquantes.xlsx

Structure réelle du classeur (confirmée sur le vrai fichier) :
- Onglet "README"    : ignoré.
- Onglet "EN_Common" :
    Ligne 0 : bannière (à ignorer).
    Ligne 1 : entêtes — Key | Page / Section | Element Type | Current English Text | Notes.
    Lignes de section : col 0 = label de section, cols 1-3 = None → à ignorer.
    Lignes de données : Key (col 0) + valeur (col 3 = "Current English Text").
    - BugType_Entry_XX  → BUG_TYPES [{id, text}].
    - Toutes les autres → EN_COMMON, avec renommage via EN_COMMON_KEY_MAP.
- Onglets de langue (ex. "fr-FR") :
    Ligne 0 : bannière emoji (à ignorer).
    Ligne 1 : entêtes — Key | EN Reference / Description | Current Text (<Langue>) | Notes.
    Colonne valeur = celle dont le header commence par "Current Text".
    Lignes de section (col 0 = bannière "PAGE …", cols 1-3 = None) → ignorées.
    Clés traitées : lang.meta.flagEmoji → flag, lang.meta.nativeName, lang.meta.region,
    global.mainHeading → demonym (extraction via regex), Glossary_Entry_XX,
    proofreading.qN.referenceEN / target, writingPrompt.q9 → q9,
    wordCount.suffixLabel → volumeLabel (optionnel).
"""

import argparse
import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook, Workbook
except ImportError:
    sys.exit("openpyxl manquant : pip install openpyxl --break-system-packages")


# --------------------------------------------------------------------------- #
# Renommage des clés EN_Common (Excel key → JSON key attendu par le JS)
# --------------------------------------------------------------------------- #

EN_COMMON_KEY_MAP = {
    "global.mainHeading.default":       "global.mainHeading",
    "hardlock.multiTab.titleDefault":   "security.defaultTitle",
    "hardlock.multiTab.bodyDefault":    "security.defaultReason",
    "hardlock.duplicateSession.title":  "security.duplicateTitle",
    "hardlock.duplicateSession.body":   "security.duplicateReason",
    "hardlock.mobileBlock.title":       "security.mobileTitle",
    "hardlock.mobileBlock.body":        "security.mobileReason",
    "alert.printScreen":                "security.printRestrictedAlert",
    "clipboard.screenshotBlockedText":  "security.screenshotClipboardMsg",
    "clipboard.copyBlockedText":        "security.copyRestrictedMsg",
    "clipboard.pasteBlockedText":       "security.pasteRestrictedMsg",
    "shield.title":                     "shield.protectedLabel",
    "shield.instructions":              "shield.instructionLabel",
    "button.resetQA":                   "buttons.resetQA",
    "button.themeToggle.dark":          "buttons.darkMode",
    "button.themeToggle.light":         "buttons.lightMode",
    "confirm.resetQA":                  "qaReset.confirmPrompt",
    "page1.tip":                        "page1.tipAlert",
    "page2.imageEnlargeHint":           "common.enlargeHint",
    "page3.glossary.mergedHeader":      "page3.glossary.sourceColumnHeader",
    "page3.button.previous":            "page3.button.prev",
    "page4.button.previous":            "page4.button.prev",
    "modal.text.default":               "modal.defaultText",
    "modal.text.warningPrefix":         "modal.warningPrefix",
    "modal.text.warningFieldsIntro":    "modal.warningBody",
    "modal.text.warningConfirm":        "modal.warningConfirm",
    "page5.thanksTitle":                "page5.thankYouTitle",
    "page5.thanksBody":                 "page5.thankYouBody",
    "page5.languageBoxPrefix":          "page5.languageMetaPrefix",
    "feedback.title":                   "page5.feedback.heading",
    "feedback.intro":                   "page5.feedback.intro",
    "feedback.option.time":             "page5.feedback.option.timeIssue",
    "feedback.option.bug":              "page5.feedback.option.functionalIssue",
    "feedback.option.unclear":          "page5.feedback.option.unclearInstructions",
    "feedback.option.other":            "page5.feedback.option.other",
    "feedback.placeholder":             "page5.feedback.textareaPlaceholder",
    "feedback.button.submit":           "page5.feedback.submitButton",
    "feedback.successMessage":          "page5.feedback.successMessage",
    "feedback.alert.empty":             "page5.feedback.alert.noSelection",
    "feedback.alert.sendError":         "page5.feedback.alert.sendError",
}

# Clés Excel à ignorer entièrement (bannières de note ou valeur toujours null).
EN_COMMON_SKIP = {
    "noscript.title",
    "noscript.body",
    "page1.languageGrid.note",
    "page2.sidebar.bugTypesBody.moved",
    "page3.referenceText.moved",
}


REQUIRED_LANGUAGE_KEYS = {
    "flag", "nativeName", "region", "demonym",
    "referenceEN.q5", "referenceEN.q6", "referenceEN.q7",
    "target.q5", "target.q6", "target.q7", "q9",
}

_GLOSSARY_ENTRY_RE = re.compile(r'^Glossary_Entry_\d+$')
_PROOFREAD_RE = re.compile(
    r'^proofreading\.(q[567])\.(referenceEN|target)$'
)


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #

def _find_header_row(rows):
    """Renvoie l'index de la ligne d'entêtes (celle dont col 0 == 'Key')."""
    for i, row in enumerate(rows):
        if row and str(row[0]).strip() == "Key":
            return i
    raise ValueError("Ligne d'entêtes 'Key' introuvable.")


def _col_index(header_row, prefix):
    """Trouve l'index de la colonne dont le header commence par `prefix`."""
    for i, cell in enumerate(header_row):
        if cell and str(cell).strip().startswith(prefix):
            return i
    raise ValueError(f"Colonne avec préfixe '{prefix}' introuvable dans : {header_row}")


def _is_section_row(row):
    """Vrai si c'est une ligne de section (col 0 non vide, tout le reste vide)."""
    if not row or row[0] is None:
        return False
    return all(c is None for c in row[1:4] if len(row) > 1)


def extract_demonym(heading: str) -> str:
    """'LQA French Tester Assessment' → 'French'."""
    m = re.match(r'^LQA\s+(.+?)\s+Tester\s+Assessment$', heading.strip())
    return m.group(1).strip() if m else ""


def extract_en_term(ref: str) -> str:
    """'EN: Orb' → 'Orb'."""
    if not ref:
        return ""
    m = re.match(r'^EN:\s*(.+)$', str(ref).strip())
    return m.group(1).strip() if m else str(ref).strip()


# --------------------------------------------------------------------------- #
# Parsing
# --------------------------------------------------------------------------- #

def parse_en_common(ws):
    """Lit EN_Common → (EN_COMMON dict, BUG_TYPES list)."""
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx = _find_header_row(rows)
    header = rows[hdr_idx]
    val_col = _col_index(header, "Current English Text")
    key_col = 0

    en_common = {}
    bug_types = []

    for row in rows[hdr_idx + 1:]:
        if not row or row[key_col] is None:
            continue
        raw_key = str(row[key_col]).strip()
        if not raw_key:
            continue
        if _is_section_row(row):
            continue
        if raw_key in EN_COMMON_SKIP:
            continue

        value = row[val_col] if len(row) > val_col and row[val_col] is not None else None
        if value is None:
            continue
        value = str(value)

        if raw_key.startswith("BugType_Entry_"):
            bug_types.append({"id": raw_key, "text": value})
        else:
            json_key = EN_COMMON_KEY_MAP.get(raw_key, raw_key)
            en_common[json_key] = value

    bug_types.sort(key=lambda b: b["id"])
    return en_common, bug_types


def parse_language_sheet(ws):
    """Lit un onglet de langue → dict (métadonnées + glossaire + textes)."""
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx = _find_header_row(rows)
    header = rows[hdr_idx]
    val_col = _col_index(header, "Current Text")
    en_ref_col = _col_index(header, "EN Reference")

    pack = {}
    glossary = []

    for row in rows[hdr_idx + 1:]:
        if not row or row[0] is None:
            continue
        raw_key = str(row[0]).strip()
        if not raw_key:
            continue
        if _is_section_row(row):
            continue

        raw_value = row[val_col] if len(row) > val_col else None
        value = str(raw_value).strip() if raw_value is not None else None
        en_ref = row[en_ref_col] if len(row) > en_ref_col else None

        # Clés méta — carte langue
        if raw_key == "lang.meta.flagEmoji":
            pack["flag"] = value or ""
        elif raw_key == "lang.meta.nativeName":
            pack["nativeName"] = value or ""
        elif raw_key == "lang.meta.region":
            pack["region"] = value or ""
        elif raw_key == "lang.meta.code":
            pass  # redondant avec le nom de l'onglet
        elif raw_key == "global.mainHeading":
            pack["demonym"] = extract_demonym(value or "")

        # Entrées glossaire
        elif _GLOSSARY_ENTRY_RE.match(raw_key):
            glossary.append({
                "id": raw_key,
                "en": extract_en_term(str(en_ref) if en_ref else ""),
                "target": value or "",
            })

        # Proofreading Q5/Q6/Q7
        elif (m := _PROOFREAD_RE.match(raw_key)):
            q, field = m.group(1), m.group(2)
            if field == "referenceEN":
                pack.setdefault("referenceEN", {})[q] = value or ""
            else:
                pack.setdefault("target", {})[q] = value or ""

        # Q9 writing prompt
        elif raw_key == "writingPrompt.q9":
            pack["q9"] = value or ""

        # Suffixe de décompte de mots (optionnel, certaines langues seulement)
        elif raw_key == "wordCount.suffixLabel":
            if value:
                pack["volumeLabel"] = value

        # Notes et rappels → ignorer
        elif raw_key in ("proofreading.bracketVariables.note", "Glossary_GeneralNote"):
            pass

    glossary.sort(key=lambda g: g["id"])

    missing = REQUIRED_LANGUAGE_KEYS - {
        k for k in REQUIRED_LANGUAGE_KEYS
        if (k in ("flag", "nativeName", "region", "demonym", "q9") and pack.get(k))
        or (k.startswith("referenceEN.") and pack.get("referenceEN", {}).get(k.split(".")[1]))
        or (k.startswith("target.") and pack.get("target", {}).get(k.split(".")[1]))
    }
    if missing:
        print(f"[AVERTISSEMENT] Onglet '{ws.title}' : cles manquantes -> {sorted(missing)}",
              file=sys.stderr)
    if len(glossary) != 10:
        print(f"[AVERTISSEMENT] Onglet '{ws.title}' : {len(glossary)} entrées glossaire "
              f"(10 attendues).", file=sys.stderr)

    result = {
        "flag":       pack.get("flag", ""),
        "nativeName": pack.get("nativeName", ""),
        "region":     pack.get("region", ""),
        "demonym":    pack.get("demonym", ""),
        "glossary":   glossary,
        "referenceEN": pack.get("referenceEN", {"q5": "", "q6": "", "q7": ""}),
        "target":      pack.get("target",      {"q5": "", "q6": "", "q7": ""}),
        "q9":         pack.get("q9", ""),
    }
    if "volumeLabel" in pack:
        result["volumeLabel"] = pack["volumeLabel"]
    return result


def build_data_from_workbook(xlsx_path: Path) -> dict:
    wb = load_workbook(xlsx_path, data_only=True)
    sheet_names = [s for s in wb.sheetnames if s not in ("README", "EN_Common")]

    en_common, bug_types = parse_en_common(wb["EN_Common"])

    # GLOSSARY_TERMS_EN : termes anglais extraits du premier onglet langue.
    glossary_terms_en = []
    if sheet_names:
        first_lang = parse_language_sheet(wb[sheet_names[0]])
        glossary_terms_en = [entry["en"] for entry in first_lang["glossary"]]

    lqa_languages = {}
    for code in sheet_names:
        pack = parse_language_sheet(wb[code])
        lqa_languages[code] = pack

    return {
        "EN_COMMON":         en_common,
        "BUG_TYPES":         bug_types,
        "GLOSSARY_TERMS_EN": glossary_terms_en,
        "LANG_GRID_ORDER":   sheet_names,
        "LQA_LANGUAGES":     lqa_languages,
    }


# --------------------------------------------------------------------------- #
# Volet 3 — détection des clés orphelines
# --------------------------------------------------------------------------- #

KEY_REFERENCE_PATTERNS = [
    re.compile(r'data-i18n(?:-placeholder|-alt)?="([^"]+)"'),
    re.compile(r"EN_COMMON\[['\"]([^'\"]+)['\"]\]"),
]


def known_keys(data: dict) -> set:
    keys = set(data["EN_COMMON"].keys())
    for pack in data["LQA_LANGUAGES"].values():
        keys |= {f"referenceEN.{q}" for q in pack["referenceEN"]}
        keys |= {f"target.{q}" for q in pack["target"]}
    return keys


def find_referenced_keys(html_path: Path) -> set:
    html = html_path.read_text(encoding="utf-8")
    found = set()
    for pattern in KEY_REFERENCE_PATTERNS:
        found |= set(pattern.findall(html))
    return found


def write_missing_keys_workbook(missing_keys: set, out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cles_manquantes"
    ws.append(["Clé", "Valeur"])
    for key in sorted(missing_keys):
        ws.append([key, ""])
    wb.save(out_path)


# --------------------------------------------------------------------------- #
# Orchestration
# --------------------------------------------------------------------------- #

def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx",        default="l18n/lqa-textes-i18n.xlsx")
    parser.add_argument("--html",        default="index_Clau_0-5.html")
    parser.add_argument("--out",         default="l18n/lqa-data.json")
    parser.add_argument("--missing-out", default="l18n/lqa-cles-manquantes.xlsx")
    args = parser.parse_args()

    xlsx_path      = Path(args.xlsx)
    html_path      = Path(args.html)
    out_path       = Path(args.out)
    missing_out    = Path(args.missing_out)

    if not xlsx_path.exists():
        sys.exit(f"Classeur introuvable : {xlsx_path}")

    data = build_data_from_workbook(xlsx_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"OK -> {out_path}  "
          f"({len(data['EN_COMMON'])} cles EN_Common, "
          f"{len(data['LQA_LANGUAGES'])} langues, "
          f"{len(data['BUG_TYPES'])} bug types, "
          f"{len(data['GLOSSARY_TERMS_EN'])} termes glossaire EN)")

    if html_path.exists():
        referenced = find_referenced_keys(html_path)
        missing = referenced - known_keys(data)
        if missing:
            write_missing_keys_workbook(missing, missing_out)
            print(f"[ATTENTION] {len(missing)} cle(s) referencee(s) dans {html_path.name} "
                  f"absente(s) du classeur -> voir {missing_out}", file=sys.stderr)
            for k in sorted(missing):
                print(f"  - {k}", file=sys.stderr)
            sys.exit(1)
        print("OK -> aucune cle orpheline detectee.")
    else:
        print(f"[INFO] {html_path} introuvable, détection des clés orphelines ignorée.")


if __name__ == "__main__":
    main()
